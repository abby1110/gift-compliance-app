"""
常態滿額禮合規分析工具
Streamlit App — 跨品牌通用版
"""
import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core import (
    load_inventory, load_gifts, load_tiers,
    load_rules, run_analysis, suggest_tier, get_depth_rate,
)
from brand_config import BRAND_CONFIGS

# ── 頁面設定 ──────────────────────────────────────────────
st.set_page_config(
    page_title="常態滿額禮合規分析",
    page_icon="🎁",
    layout="wide",
)

st.title("🎁 常態滿額禮合規分析工具")
st.caption("上傳四份資料，自動計算各門檻贈品折扣深度合規性，並給出替換建議")

# ── Sidebar：品牌與分析設定 ───────────────────────────────
with st.sidebar:
    st.header("⚙️ 分析設定")

    brand = st.selectbox("品牌", list(BRAND_CONFIGS.keys()))
    cfg = BRAND_CONFIGS[brand]

    st.divider()

    platform = st.selectbox("平台（規範表工作表）", ["蝦皮", "官網", "momo", "HKTVmall"])
    scale = st.selectbox("檔期規模", ["常態", "［S］SBD、雙11、618", "［A］雙12、寵物節", "［B］DD、18、25"])

    st.divider()

    market = st.radio("分析市場", ["台灣 (NTD)", "港澳 (HKD)"])
    is_tw = market.startswith("台灣")
    market_key = "tw" if is_tw else "hk"
    currency = "ntd" if is_tw else "hkd"
    cur_symbol = "NTD" if is_tw else "HKD"

    st.divider()
    st.caption("需要同時分析兩個市場？先跑台灣，下載後再切換港澳再跑一次")

# ── 主區域：檔案上傳 ─────────────────────────────────────
st.subheader("📂 上傳資料")

col1, col2, col3, col4 = st.columns(4)

with col1:
    st.markdown("**① 庫存表**")
    st.caption("即時庫存，當天下載")
    inv_file = st.file_uploader("庫存表", type=["xlsx"], key="inv",
                                label_visibility="collapsed")

with col2:
    st.markdown("**② 折扣深度規範表**")
    st.caption("電商活動折扣規範")
    rules_file = st.file_uploader("規範表", type=["xlsx"], key="rules",
                                  label_visibility="collapsed")

with col3:
    st.markdown("**③ 贈品資訊表**")
    st.caption("貨號、類型、台幣、港幣價格")
    gifts_file = st.file_uploader("贈品資訊", type=["xlsx"], key="gifts",
                                  label_visibility="collapsed")

with col4:
    st.markdown("**④ 門檻設定表**")
    st.caption("台灣與港澳各檻金額")
    tiers_file = st.file_uploader("門檻設定", type=["xlsx"], key="tiers",
                                  label_visibility="collapsed")

# ── 說明：贈品資訊表格式 ──────────────────────────────────
with st.expander("📋 贈品資訊表格式說明", expanded=False):
    st.markdown(f"""
    **必要欄位（欄位名稱可在品牌設定中對應）：**

    | 欄位 | 說明 | 範例 |
    |------|------|------|
    | 商品選項貨號 | 和庫存表對應的 key | `牛排貓抓板` |
    | 類型 | 正商品 或 特規品 | `特規品` |
    | 台幣價格 | 正商品=折扣售價；特規品=成本NTD | `46.0` |
    | 港幣價格 | 正商品=HKD售價；特規品=成本HKD | `11.5` |
    | 狀態 | 啟用中 / 贈完 / 下架 | `啟用中` |
    | 是否指定 | 第一檻指定品免規範檢查 | `TRUE` |
    | 台灣門檻 | (選填) 目前掛的台灣門檻金額 | `1200` |
    | 港澳門檻 | (選填) 目前掛的港澳門檻金額 | `400` |
    """)

# ── 執行分析 ──────────────────────────────────────────────
all_uploaded = all([inv_file, rules_file, gifts_file, tiers_file])

if all_uploaded:
    run_btn = st.button("🚀 開始分析", type="primary", use_container_width=True)
else:
    missing = sum([inv_file is None, rules_file is None,
                   gifts_file is None, tiers_file is None])
    st.info(f"還需上傳 {missing} 份資料才能開始分析")
    run_btn = False

if run_btn:
    with st.spinner("讀取資料中..."):
        try:
            inventory = load_inventory(inv_file, cfg)
            rules_df  = load_rules(rules_file)
            gifts_df  = load_gifts(gifts_file, cfg)
            tiers_dict = load_tiers(tiers_file, cfg)
        except Exception as e:
            st.error(f"讀取資料失敗：{e}")
            st.stop()

    with st.spinner("計算合規性..."):
        result_df = run_analysis(
            gifts_df, inventory, tiers_dict, rules_df,
            platform, scale, market_key, currency
        )

    # ── 摘要統計 ──────────────────────────────────────────
    st.divider()
    st.subheader("📊 分析結果")

    total = len(result_df)
    over  = (result_df["適用性"] == "否,太高").sum()
    low   = (result_df["適用性"] == "否,太低").sum()
    ok    = (result_df["適用性"] == "是").sum()
    pending = total - over - low - ok

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("全部贈品", total)
    c2.metric("🔴 超標需汰換", int(over))
    c3.metric("🔵 價值偏低", int(low))
    c4.metric("✅ 合規", int(ok))
    c5.metric("⏳ 待補充", int(pending))

    # ── 顯示結果表格 ──────────────────────────────────────
    st.divider()

    # 分頁顯示
    tab_over, tab_low, tab_ok, tab_all = st.tabs([
        f"🔴 超標 ({over})",
        f"🔵 偏低 ({low})",
        f"✅ 合規 ({ok})",
        f"全部 ({total})",
    ])

    display_cols = [
        "贈品貨號", "類型", "門檻", "贈品價值",
        "規範折扣金額", "規範50%", "贈品深度", "規範深度",
        "適用性", ">20%", ">50%", "狀態", "庫存", "建議門檻"
    ]
    pct_cols = ["贈品深度", "規範深度"]

    def fmt_df(df):
        d = df[display_cols].copy()
        for col in pct_cols:
            if col in d.columns:
                d[col] = d[col].apply(
                    lambda x: f"{x*100:.2f}%" if pd.notna(x) else "—"
                )
        return d.fillna("—")

    def color_rows(row):
        appl = row.get("適用性", "")
        if appl == "否,太高":
            return ["background-color: #FFCCCC"] * len(row)
        elif appl == "否,太低":
            return ["background-color: #E6EFF6"] * len(row)
        elif appl == "是":
            return ["background-color: #E5F0EB"] * len(row)
        return [""] * len(row)

    with tab_over:
        df_over = result_df[result_df["適用性"] == "否,太高"]
        if df_over.empty:
            st.success("沒有超標贈品！")
        else:
            st.dataframe(
                fmt_df(df_over).style.apply(color_rows, axis=1),
                use_container_width=True, hide_index=True
            )

    with tab_low:
        df_low = result_df[result_df["適用性"] == "否,太低"]
        if df_low.empty:
            st.info("沒有偏低贈品")
        else:
            st.dataframe(
                fmt_df(df_low).style.apply(color_rows, axis=1),
                use_container_width=True, hide_index=True
            )

    with tab_ok:
        df_ok = result_df[result_df["適用性"] == "是"]
        if df_ok.empty:
            st.warning("目前無合規贈品")
        else:
            st.dataframe(
                fmt_df(df_ok).style.apply(color_rows, axis=1),
                use_container_width=True, hide_index=True
            )

    with tab_all:
        st.dataframe(
            fmt_df(result_df).style.apply(color_rows, axis=1),
            use_container_width=True, hide_index=True
        )

    # ── 匯出 Excel ────────────────────────────────────────
    st.divider()
    st.subheader("📥 匯出報告")

    def build_excel(df, brand_name, market_label, cur):
        wb = Workbook()
        ws = wb.active
        ws.title = f"{market_label}合規分析"

        RED_F    = PatternFill("solid", start_color="FFCCCC")
        AMBER_F  = PatternFill("solid", start_color="FFE699")
        GREEN_F  = PatternFill("solid", start_color="C6EFCE")
        BLUE_F   = PatternFill("solid", start_color="BDD7EE")
        HEADER_F = PatternFill("solid", start_color="D9D9D9")

        RED_FONT   = Font(name="Arial", color="9C0006", bold=True,  size=10)
        AMBER_FONT = Font(name="Arial", color="7D5A00", bold=True,  size=10)
        GREEN_FONT = Font(name="Arial", color="276221", bold=True,  size=10)
        BLUE_FONT  = Font(name="Arial", color="1F497D", bold=True,  size=10)
        BOLD_FONT  = Font(name="Arial", bold=True, size=10)
        NORM_FONT  = Font(name="Arial", size=10)

        thin   = Side(style="thin", color="BFBFBF")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
        CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT   = Alignment(horizontal="left", vertical="center")

        HEADERS = [
            "贈品貨號", "類型", "門檻", "狀態", "庫存",
            "贈品價值", f"規範折扣金額({cur})", f"規範50%({cur})",
            "贈品深度", "規範深度", "適用性", ">20%", ">50%", "建議門檻"
        ]
        WIDTHS = [20, 10, 10, 10, 10, 10, 14, 12, 10, 10, 12, 10, 10, 12]

        for c, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = BOLD_FONT
            cell.fill = HEADER_F
            cell.alignment = CENTER
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = 28

        cols_map = {h: i+1 for i, h in enumerate(HEADERS)}

        for row_i, (_, row) in enumerate(df.iterrows(), 2):
            appl = str(row.get("適用性", ""))
            for c, h in enumerate(HEADERS, 1):
                key_map = {
                    "贈品貨號": "贈品貨號", "類型": "類型", "門檻": "門檻",
                    "狀態": "狀態", "庫存": "庫存", "贈品價值": "贈品價值",
                    f"規範折扣金額({cur})": "規範折扣金額",
                    f"規範50%({cur})": "規範50%",
                    "贈品深度": "贈品深度", "規範深度": "規範深度",
                    "適用性": "適用性", ">20%": ">20%", ">50%": ">50%",
                    "建議門檻": "建議門檻",
                }
                val = row.get(key_map.get(h, h), "")
                cell = ws.cell(row=row_i, column=c, value=val)
                cell.font = NORM_FONT
                cell.border = BORDER
                cell.alignment = LEFT if c == 1 else CENTER
                if h in ("贈品深度", "規範深度") and isinstance(val, float):
                    cell.number_format = "0.00%"
                elif h in ("贈品價值", f"規範折扣金額({cur})", f"規範50%({cur})"):
                    cell.number_format = "#,##0.0"

            # 適用性上色
            ac = ws.cell(row=row_i, column=cols_map["適用性"])
            if appl == "否,太高":     ac.fill = RED_F;   ac.font = RED_FONT
            elif appl == "否,太低":   ac.fill = BLUE_F;  ac.font = BLUE_FONT
            elif appl == "是":        ac.fill = GREEN_F; ac.font = GREEN_FONT
            elif "需調整" in str(row.get(">20%","")):
                ws.cell(row=row_i, column=cols_map[">20%"]).fill = RED_F
                ws.cell(row=row_i, column=cols_map[">20%"]).font = RED_FONT
            ws.row_dimensions[row_i].height = 18

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    excel_buf = build_excel(result_df, brand, market, cur_symbol)

    import datetime
    today = datetime.date.today().strftime("%Y%m%d")
    filename = f"{brand}_{market_key.upper()}_{today}_合規分析.xlsx"

    st.download_button(
        label="⬇️ 下載 Excel 報告",
        data=excel_buf,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.caption(f"報告產生時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
